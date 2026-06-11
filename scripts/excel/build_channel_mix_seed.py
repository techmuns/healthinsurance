#!/usr/bin/env python3
"""
Channel Mix HISTORY SEED — from Neha's Channel Mix workbook.

Neha's instruction (2026-06-11): the Channel Mix tab must match her uploaded
sheet exactly and every cell must be filled and kept automated. Her sheet IS
the official IRDAI NL-36/NL-40 business-acquisition data (verified: the Care
FY19 and FY24 forms reproduce her mix, avg-premium, agents-GWP and policies
figures to the 4th decimal), so it seeds the history while the automated
NL-form parser (scripts/ingest/ingest-distribution.ts) supersedes each cell
as the official disclosure for that period is staged and parsed.

This script reads the committed workbook
(data/uploads/channel-mix-seed-workbook.xlsx) cell-by-cell through the SAME
schema-map bindings the template uses, and writes a seed value file.

Honesty contract (same as the Industry Growth seed):
  * Every seeded value is labelled as coming from Neha's workbook (clearly
    NOT an official source) and carries the committed file + repo URL.
  * Seeds enter the value store at rank 8 — every official source (NL-form
    parses, filings, snapshots) outranks and silently supersedes a seed; a
    seed can never flag a conflict against an official value.
  * Blank cells stay blank. Missing is never zero.
  * UNLIKE the Industry Growth seed, printed zeros are KEPT: in this sheet a
    zero is a real figure (e.g. 0% commission on Direct Business — there is
    no intermediary to pay), not a not-applicable era. Tiny negatives
    (premium refund adjustments, as printed) are kept too.

Run:  python3 scripts/excel/build_channel_mix_seed.py
Out:  data/source-map/channel-mix-seed.json
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "data" / "uploads" / "channel-mix-seed-workbook.xlsx"
SCHEMA = REPO / "schema-map.json"
OUT = REPO / "data" / "source-map" / "channel-mix-seed.json"

SOURCE_NAME = ("Neha's portfolio-review workbook — Channel Mix sheet (IRDAI NL-36/NL-40 basis), "
               "user-provided 2026-06-11 as the historical seed")
SOURCE_FILE = "data/uploads/channel-mix-seed-workbook.xlsx"
SOURCE_URL = "https://github.com/techmuns/HealthInsurance/blob/main/data/uploads/channel-mix-seed-workbook.xlsx"
NOTE = ("Historical seed per Neha (2026-06-11). Superseded automatically by any official "
        "source for the same cell (the IRDAI NL-36/NL-40 business-acquisition parse outranks it). "
        "Basis: channel premium share of Total(A), up-to-period column; premium metrics (not profit).")


def main() -> None:
    schema = json.loads(SCHEMA.read_text())
    sheet = next(s for s in schema["sheets"] if s["sheet"] == "Channel Mix")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]  # her export carries the grid on its only sheet

    entries: dict[str, dict] = {}
    skipped_blank = kept_zero = 0
    for b in sheet["bindings"]:
        if not b.get("fillable") or b.get("cell_kind") != "input":
            continue
        v = ws[b["cell"]].value
        if v is None or not isinstance(v, (int, float)):
            skipped_blank += 1
            continue
        if v == 0:
            kept_zero += 1  # genuine printed zero (e.g. direct business commission)
        key = f"{b['entity']}::{b['metric']}::{b['period']}"
        entries[key] = {
            "value": round(float(v), 8),
            "unit": b.get("unit") or "ratio",
            "cell": b["cell"],
            "source_name": SOURCE_NAME,
            "source_file": SOURCE_FILE,
            "source_url": SOURCE_URL,
            "note": NOTE,
        }

    OUT.write_text(json.dumps({
        "_meta": {
            "artifact": "channel-mix-seed",
            "description": ("Historical seed values for the Channel Mix sheet, read from Neha's workbook "
                            "via the schema-map cell contract. Rank-8 in the value store: every official "
                            "source supersedes a seed. Printed zeros are genuine values here (kept)."),
            "workbook": SOURCE_FILE,
            "generated_by": "scripts/excel/build_channel_mix_seed.py",
        },
        "values": entries,
    }, indent=2, ensure_ascii=False) + "\n")
    print(f"seeded {len(entries)} cells from the workbook "
          f"({skipped_blank} blank stayed empty — missing ≠ zero; {kept_zero} genuine zeros kept)")


if __name__ == "__main__":
    main()
