#!/usr/bin/env python3
"""
Industry Growth HISTORY SEED — from Neha's original workbook.

Neha's instruction (2026-06-10): the FY25/FY26 columns are filled from the
official GI Council reports; seed the HISTORICAL years from the Industry
Growth sheet of the workbook she provided ("sent to AI Team" — the Capital
IQ-era numbers). This script reads that workbook (committed at
data/uploads/industry-growth-seed-workbook.xlsx) cell-by-cell through the
SAME schema-map bindings the template uses, and writes a seed value file.

Honesty contract:
  * Every seeded value is labelled as coming from Neha's workbook (clearly
    NOT an official source) and carries the committed file + repo URL.
  * Seeds enter the value store at rank 8 — every official source (GI
    Council reports, filings, snapshots) outranks and silently supersedes a
    seed; a seed can never flag a conflict against an official value.
  * Blank cells stay blank. Missing is never zero.

Run:  python3 scripts/excel/build_industry_growth_seed.py
Out:  data/source-map/industry-growth-seed.json
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
WORKBOOK = REPO / "data" / "uploads" / "industry-growth-seed-workbook.xlsx"
SCHEMA = REPO / "schema-map.json"
OUT = REPO / "data" / "source-map" / "industry-growth-seed.json"

SOURCE_NAME = ("Neha's portfolio-review workbook (Capital IQ era) — Industry Growth sheet, "
               "user-provided 2026-06-10 as the historical seed")
SOURCE_FILE = "data/uploads/industry-growth-seed-workbook.xlsx"
SOURCE_URL = "https://github.com/techmuns/HealthInsurance/blob/main/data/uploads/industry-growth-seed-workbook.xlsx"
NOTE = ("Historical seed per Neha (2026-06-10). Superseded automatically by any official "
        "source for the same cell (GI Council segment reports outrank it).")


def main() -> None:
    schema = json.loads(SCHEMA.read_text())
    sheet = next(s for s in schema["sheets"] if s["sheet"] == "Industry Growth")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]  # her export carries the grid on its only sheet

    entries: dict[str, dict] = {}
    skipped_blank = skipped_zero = 0
    for b in sheet["bindings"]:
        if not b.get("fillable") or b.get("cell_kind") != "input":
            continue
        v = ws[b["cell"]].value
        if v is None or not isinstance(v, (int, float)):
            skipped_blank += 1
            continue
        if v == 0:
            # The workbook prints 0 for not-applicable eras (insurer not yet
            # licensed / already merged away). That is absence, not a premium
            # of zero — never seeded. Missing ≠ zero.
            skipped_zero += 1
            continue
        key = f"{b['entity']}::{b['metric']}::{b['period']}"
        entries[key] = {
            "value": round(float(v), 2),
            "unit": b.get("unit") or "INR_cr",
            "cell": b["cell"],
            "source_name": SOURCE_NAME,
            "source_file": SOURCE_FILE,
            "source_url": SOURCE_URL,
            "note": NOTE,
        }

    OUT.write_text(json.dumps({
        "_meta": {
            "artifact": "industry-growth-seed",
            "description": "Historical seed values for the Industry Growth sheet, read from Neha's workbook via the schema-map cell contract. Rank-8 in the value store: every official source supersedes a seed.",
            "workbook": SOURCE_FILE,
            "generated_by": "scripts/excel/build_industry_growth_seed.py",
        },
        "values": entries,
    }, indent=2, ensure_ascii=False) + "\n")
    print(f"seeded {len(entries)} cells from the workbook "
          f"({skipped_blank} blank + {skipped_zero} not-applicable zeros stayed empty — missing ≠ zero)")


if __name__ == "__main__":
    main()
