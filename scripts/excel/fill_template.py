#!/usr/bin/env python3
"""
Phase 5 - Template filler (Chunk 2B).

Reads ``schema-map.json`` (the cell contract) and ``data/processed/excel-values.json``
(the normalized, source-backed value store) and writes a filled copy of the
workbook, preserving the original formatting and every Excel formula.

Honesty model (per project policy + the governing charter):
* The source template is treated as a *layout only*. We do NOT keep its original
  numbers. Each fillable cell is either (a) written from a source-backed value
  that passed every gate, or (b) left blank and recorded on **Missing Data** or
  **Blocked Data**. Missing != zero.
* A value is FILLED only when it has a normalized value AND is not flagged
  ``conflict_needs_review``. Source-conflicting values are kept in the store but
  NOT filled - they land on Blocked Data so a reviewer can adjudicate.
* Three sheets are appended: **Source Audit** (every filled value, full
  documentary provenance so the dashboard can click through to the source),
  **Missing Data** (cells with no source value yet), and **Blocked Data**
  (extracted-but-withheld values, categorized: parser_failed, mangled_extraction,
  period_unclear, unit_unclear, basis_unclear, scope_unclear, low_confidence_ppt,
  source_conflict).

Usage:
    python3 scripts/excel/fill_template.py [template.xlsx] [output.xlsx]
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parents[2]
SCHEMA = REPO / "schema-map.json"
VALUES = REPO / "data" / "processed" / "excel-values.json"
HELD_BACK = REPO / "data" / "processed" / "excel-held-back.json"
FILINGS = REPO / "src" / "data" / "snapshots" / "company-filings-snapshot.json"
TEMPLATE = REPO / "templates" / "niva-bupa-portfolio-review.xlsx"
OUT = REPO / "output" / "Niva_Bupa_portfolio_review__filled.xlsx"

AUDIT_SHEET = "Source Audit"
MISSING_SHEET = "Missing Data"
BLOCKED_SHEET = "Blocked Data"

MISSING_REASON = {
    "available": "Source supported but value not yet fetched - run the ingestion job (needs internet egress).",
    "partial": "Partially available from official disclosures - period/coverage gap.",
    "backup": "No official equivalent; backup aggregator (Screener/Trendlyne) returned no value.",
    "computed": "Computed in-sheet by an Excel formula - no fetch required.",
    "narrative": "Editorial summary from transcripts - populated by a reviewer, not a numeric fetch.",
    "excluded_from_core": "Outside the source-backed core dataset (curated news).",
}

PPT_TYPES = {"investor_presentation", "quarterly_ppt", "investor_ppt"}


def header(ws, titles, fill="1F4E5F"):
    bold = Font(bold=True, color="FFFFFF")
    pat = PatternFill("solid", fgColor=fill)
    for c, t in enumerate(titles, start=1):
        cell = ws.cell(1, c, t)
        cell.font = bold
        cell.fill = pat
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load_json(path, default):
    return json.loads(path.read_text()) if path.exists() else default


def blocked_category(rec):
    """Map a blocked company-filings record to one separated category."""
    dt = rec.get("document_type")
    es = rec.get("extraction_status")
    reason = f"{rec.get('sanity_reason') or ''} {rec.get('parser_notes') or ''}".lower()
    if dt in PPT_TYPES:
        return "low_confidence_ppt"
    if es == "mangled" or "fused-column" in reason or "mangled" in reason:
        return "mangled_extraction"
    if es in ("parser_failed", "no_metrics_found"):
        return "parser_failed"
    if "period" in reason and "unclear" in reason:
        return "period_unclear"
    if "unit" in reason and "unclear" in reason:
        return "unit_unclear"
    # Chunk 2C-A: the column-aware NL-form parser withholds a value when the
    # table columns cannot be lined up with certainty (e.g. a point-in-time
    # metric disagrees across its standalone/YTD columns).
    if "column" in reason or "alignment" in reason:
        return "column_unclear"
    return "needs_review"


def main(template_path: Path, out_path: Path) -> None:
    schema = json.loads(SCHEMA.read_text())
    store = json.loads(VALUES.read_text()) if VALUES.exists() else {}
    held = load_json(HELD_BACK, {"data": []}).get("data", [])
    filings = load_json(FILINGS, {"data": []}).get("data", [])
    # Cells where the statutory 1/n basis was selected over an adjusted ex-1/n value
    # (an alternate-basis record exists on Blocked Data) -> annotate the audit row.
    basis_selected = {(h.get("company_id"), h.get("metric"), h.get("filing_period"))
                      for h in held if h.get("hold_reason") == "basis_mismatch_ex_1n_adjusted"}

    wb = openpyxl.load_workbook(template_path)  # keep formulas
    for s in (AUDIT_SHEET, MISSING_SHEET, BLOCKED_SHEET):
        if s in wb.sheetnames:
            del wb[s]

    audit_rows, missing_rows, blocked_rows = [], [], []
    filled = skipped = conflict_blocked = 0

    for sheet in schema["sheets"]:
        name = sheet["sheet"]
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for b in sheet["bindings"]:
            if not b.get("fillable"):
                continue
            key = f"{b['entity']}::{b['metric']}::{b['period']}"
            entry = store.get(key)
            cell = b["cell"]
            has_value = bool(entry and entry.get("normalized_value") is not None)
            is_conflict = bool(entry and entry.get("conflict_status") == "conflict_needs_review")

            if has_value and not is_conflict:
                ws[cell] = entry["normalized_value"]
                filled += 1
                audit_rows.append([
                    name, cell, b["entity"], b["metric"], b["period"], b["unit"],
                    entry.get("raw_value"), entry.get("normalized_value"),
                    entry.get("transformation_used"), entry.get("source_name"),
                    entry.get("source_url"), entry.get("fetched_at"),
                    entry.get("confidence"), entry.get("source_status", "available"),
                    # --- extended documentary provenance (Chunk 2B) ---
                    entry.get("source_layer"), entry.get("document_type"),
                    entry.get("document_title"), entry.get("source_file"),
                    entry.get("filing_date"), entry.get("extraction_status"),
                    entry.get("sanity_status"), entry.get("conflict_status", "none"),
                    # --- which NL-form column the value came from (Chunk 2C-A) ---
                    entry.get("column_basis"),
                    # --- basis note: explicit per-value note (deck/annual-report
                    # basis), else the statutory-1/n selection marker ---
                    entry.get("basis_note")
                    or ("Selected statutory 1/n basis for comparability"
                        if (b["entity"], b["metric"], b["period"]) in basis_selected else ""),
                ])
            elif has_value and is_conflict:
                # Source-conflicting: keep in store, do NOT fill (charter rule).
                ws[cell] = None
                conflict_blocked += 1
                comp = (entry.get("competing_values") or [{}])[0]
                blocked_rows.append([
                    "source_conflict", b["entity"], b["metric"], b["period"],
                    entry.get("raw_value"), entry.get("normalized_value"),
                    entry.get("confidence"), entry.get("document_type") or entry.get("source_layer"),
                    entry.get("filing_date"), entry.get("source_url") or entry.get("source_file"),
                    f"conflicts with {comp.get('source_layer')} value {comp.get('normalized_value')} "
                    f"(rank {comp.get('priority_rank')}); higher-priority value kept in store, cell withheld.",
                ])
            else:
                ws[cell] = None
                skipped += 1
                status = b.get("source_status", "available")
                missing_rows.append([
                    name, cell, b["entity"], b["metric"], b["period"], b["unit"],
                    status, MISSING_REASON.get(status, "Unavailable."),
                    b.get("source_name", ""),
                    "unavailable_publicly" if status in ("backup", "excluded_from_core") else "pending_fetch",
                ])

    # --- Blocked Data: held-back (basis/scope) + blocked filings records ----
    for h in held:
        blocked_rows.append([
            h.get("hold_reason"), h.get("company_id"), h.get("metric"), h.get("filing_period"),
            h.get("raw_value"), h.get("normalized_value"), h.get("confidence"),
            h.get("document_type"), h.get("filing_date"),
            h.get("source_url") or h.get("source_file"), h.get("note"),
        ])
    for rec in filings:
        if rec.get("eligible_for_excel"):
            continue
        cat = blocked_category(rec)
        reason = rec.get("sanity_reason") or rec.get("parser_notes") or rec.get("suggested_manual_fallback") or ""
        blocked_rows.append([
            cat, rec.get("company_id"), rec.get("metric"), rec.get("filing_period"),
            rec.get("raw_value"), rec.get("normalized_value"),
            rec.get("provenance", {}).get("confidence"), rec.get("document_type"),
            rec.get("filing_date"), rec.get("source_url") or rec.get("source_file"), reason,
        ])

    # --- Source Audit sheet (extended documentary provenance) --------------
    aud = wb.create_sheet(AUDIT_SHEET)
    audit_cols = ["Sheet", "Cell", "Entity", "Metric", "Period", "Unit",
                  "Raw value", "Normalized value", "Transformation",
                  "Source name", "Source URL", "Fetched at", "Confidence", "Status",
                  "Source layer", "Document type", "Document title", "Source file",
                  "Filing date", "Extraction status", "Sanity status", "Conflict status",
                  "Column basis", "Basis note"]
    header(aud, audit_cols)
    autosize(aud, [18, 7, 14, 26, 12, 9, 13, 15, 30, 30, 44, 22, 11, 11,
                   15, 18, 30, 30, 12, 16, 12, 16, 18, 44])
    for r in audit_rows:
        aud.append(r)

    # --- Missing Data sheet ------------------------------------------------
    mis = wb.create_sheet(MISSING_SHEET)
    header(mis, ["Sheet", "Cell", "Entity", "Metric", "Period", "Unit",
                 "Source status", "Reason", "Intended source", "Marker"], fill="7A3B2E")
    autosize(mis, [18, 7, 14, 26, 12, 9, 14, 52, 40, 20])
    for r in missing_rows:
        mis.append(r)

    # --- Blocked Data sheet (extracted but withheld, categorized) ----------
    blk = wb.create_sheet(BLOCKED_SHEET)
    header(blk, ["Category", "Company", "Metric", "Period", "Raw value",
                 "Normalized value", "Confidence", "Document type / layer",
                 "Filing date", "Source (URL or file)", "Reason / note"], fill="5A4A1F")
    autosize(blk, [20, 16, 20, 11, 13, 15, 11, 22, 12, 46, 60])
    blocked_rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2])))
    for r in blocked_rows:
        blk.append(r)

    # --- Run banner on the audit sheet (top, above the freeze) ------------
    aud.insert_rows(1)
    banner = aud.cell(1, 1,
        f"Source-backed fill - generated {datetime.now(timezone.utc).isoformat()} - "
        f"{filled} cells filled from official sources, {skipped} missing, "
        f"{len(blocked_rows)} blocked/withheld (incl. {conflict_blocked} source-conflict). "
        f"Template treated as layout only; original values NOT retained. Every filled "
        f"cell is traceable to its source document below.")
    banner.font = Font(italic=True, color="555555")
    aud.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(audit_cols))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    total = filled + skipped + conflict_blocked
    pct = (100.0 * filled / total) if total else 0.0
    print(f"filled workbook written -> {out_path}")
    print(f"  fillable cells: {total}")
    print(f"  filled from official sources: {filled} ({pct:.1f}%)")
    print(f"  marked missing (Missing Data sheet): {skipped}")
    print(f"  withheld at a cell (source-conflict): {conflict_blocked}")
    print(f"  Blocked Data rows (held-back + blocked filings + conflicts): {len(blocked_rows)}")
    print(f"  audit rows: {len(audit_rows)}")


if __name__ == "__main__":
    tpl = Path(sys.argv[1]) if len(sys.argv) > 1 else TEMPLATE
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT
    if not tpl.exists():
        sys.exit(f"Template not found: {tpl}")
    main(tpl, out)
