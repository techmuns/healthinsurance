#!/usr/bin/env python3
"""
Phase 1 - Template Reader.

Reads the Niva Bupa portfolio-review workbook (treated as a SCHEMA/TEMPLATE,
never as final data) and emits ``schema-map.json``: a cell-level contract that
says, for every populated region of every sheet, what metric / entity / period
each cell represents, whether the cell is a fillable *input* or an Excel
*formula* we must leave alone, and which approved source should populate it.

Design notes
------------
* The workbook is **formula-driven**: a large share of populated cells are
  Excel calculations (totals, YoY %, segment mix %, quarter splits derived as
  "cumulative minus previous"). Only the *input* cells need source data. The
  builder classifies every cell as ``label`` / ``input`` / ``formula`` /
  ``text`` / ``empty`` by inspecting both the cached value and the raw formula,
  so the filler (Phase 5) can fill inputs and preserve formulas untouched.
* Per-sheet layout is declared in ``LAYOUTS`` below (authored from a full read
  of the file). The engine expands each declared block into concrete per-cell
  bindings. Anything the engine cannot bind still appears in the grid summary,
  so nothing is silently dropped.
* Source bindings follow the project's **official-first** policy (Neha,
  2026-06-05): IRDAI / GI Council / NSE-BSE / company public disclosures are the
  source of truth; Screener / Trendlyne / Investing are only ever a clearly
  tagged, lower-confidence *backup* for cells with no official equivalent.

This script is READ-ONLY with respect to the workbook. It never modifies it.
"""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

REPO = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE = REPO / "templates" / "niva-bupa-portfolio-review.xlsx"
OUT = REPO / "schema-map.json"


# ---------------------------------------------------------------------------
# Source bindings (official-first; aggregators are backup-only and tagged).
# ---------------------------------------------------------------------------
SOURCES = {
    "industry_premium": {
        "primary_source": "IRDAI / GI Council segment business statistics",
        "primary_url": "https://www.gicouncil.in/statistics/business-statistics/",
        "adapter": "fetch_gicouncil / fetch_irdai",
        "access": "public",
        "status": "available",
    },
    "company_premium_monthly": {
        "primary_source": "IRDAI Monthly Business Figures (non-life / health)",
        "primary_url": "https://irdai.gov.in/monthly-business-figures-non-life-insurers",
        "adapter": "fetch_irdai (monthly)",
        "access": "public",
        "status": "available",
    },
    "company_premium_quarterly": {
        "primary_source": "Company public disclosures (IRDAI NL forms) + IRDAI monthly cumulative",
        "primary_url": "https://irdai.gov.in/public-disclosures",
        "adapter": "fetch_irdai / ingest-company-disclosures",
        "access": "public",
        "status": "available",
    },
    "company_financials": {
        "primary_source": "Company public disclosures (IRDAI L/NL forms) + annual reports",
        "primary_url": "https://irdai.gov.in/public-disclosures",
        "adapter": "ingest-company-disclosures / ingest-quarterly-disclosures",
        "access": "public",
        "status": "available",
    },
    "market_quote": {
        "primary_source": "NSE (price, traded & deliverable quantity)",
        "primary_url": "https://www.nseindia.com",
        "adapter": "fetch_investing (NSE-first)",
        "backup_source": "Investing.com / Trendlyne (public)",
        "access": "public",
        "status": "available",
    },
    "market_cap": {
        "primary_source": "NSE market cap (price x shares outstanding)",
        "primary_url": "https://www.nseindia.com",
        "adapter": "fetch_investing (NSE-first)",
        "backup_source": "Screener / Trendlyne (public)",
        "access": "public",
        "status": "available",
    },
    "valuation_multiple_computed": {
        "primary_source": "Computed in-sheet from price + reported financials",
        "primary_url": "",
        "adapter": "(Excel formula - not fetched)",
        "access": "n/a",
        "status": "computed",
    },
    "valuation_history": {
        "primary_source": "NSE price history + reported EPS (computed)",
        "primary_url": "https://www.nseindia.com",
        "adapter": "fetch_investing (NSE-first)",
        "backup_source": "Screener / Trendlyne 3-yr average P/E (public, backup)",
        "access": "public",
        "status": "backup",
    },
    "shareholding": {
        "primary_source": "NSE/BSE shareholding pattern filing",
        "primary_url": "https://www.nseindia.com/companies-listing/corporate-filings-shareholding-pattern",
        "adapter": "ingest-ownership",
        "backup_source": "Trendlyne / Screener shareholding (public, backup)",
        "access": "public",
        "status": "available",
    },
    "analyst_coverage": {
        "primary_source": "Broker research aggregators (no official equivalent exists)",
        "primary_url": "https://trendlyne.com",
        "adapter": "fetch_trendlyne / ingest-moneycontrol-analyst",
        "backup_source": "Moneycontrol estimates",
        "access": "public",
        "status": "backup",
        "note": "Analyst targets are inherently aggregator-sourced; tagged low-confidence per official-first policy.",
    },
    "distribution": {
        "primary_source": "Company annual reports + IRDAI NL forms (commission)",
        "primary_url": "https://irdai.gov.in/public-disclosures",
        "adapter": "ingest-distribution",
        "access": "public",
        "status": "partial",
    },
    "management_commentary": {
        "primary_source": "Company earnings-call transcripts / investor presentations",
        "primary_url": "(per-company investor-relations pages)",
        "adapter": "ingest-management-events",
        "access": "public",
        "status": "narrative",
        "note": "Editorial summary, not a fetchable figure - drafted from transcripts, reviewed by a human.",
    },
    "sector_news": {
        "primary_source": "Financial press (ET / Moneycontrol / Reuters / Business Standard)",
        "primary_url": "",
        "adapter": "(curated; non-core)",
        "access": "public",
        "status": "excluded_from_core",
        "note": "News list is explicitly outside the source-backed core dataset.",
    },
}


# ---------------------------------------------------------------------------
# Period axes (column-letter -> (period_label, period_type)).
# ---------------------------------------------------------------------------
def fy_axis(start_col: str, start_fy: int, end_fy: int) -> dict:
    """FY columns laid left-to-right, one column per fiscal year."""
    axis = {}
    c = column_index_from_string(start_col)
    for fy in range(start_fy, end_fy + 1):
        axis[get_column_letter(c)] = (f"FY{fy:02d}", "annual")
        c += 1
    return axis


GWP_FY_AXIS = {  # FY26 GWP sheet: previous FY (C-H) then current FY (I-N)
    "C": ("Q1FY25", "quarterly"),
    "D": ("Q2FY25", "quarterly"),
    "E": ("H1FY25", "quarterly_cumulative"),
    "F": ("9MFY25", "quarterly_cumulative"),
    "G": ("Q4FY25", "quarterly"),
    "H": ("FY25", "annual"),
    "I": ("Q1FY26", "quarterly"),
    "J": ("Q2FY26", "quarterly"),
    "K": ("H1FY26", "quarterly_cumulative"),
    "L": ("9MFY26", "quarterly_cumulative"),
    "M": ("Q4FY26", "quarterly"),
    "N": ("FY26", "annual"),
}

Q1_MONTHLY_AXIS = {  # Q1'26 GWP sheet: monthly feeder
    "C": ("Apr-FY25", "monthly"),
    "D": ("May-FY25", "monthly"),
    "E": ("Jun-FY25", "monthly"),
    "F": ("Q1FY25", "quarterly"),
    "G": ("Apr-FY26", "monthly"),
    "H": ("May-FY26", "monthly"),
    "I": ("Jun-FY26", "monthly"),
    "J": ("Q1FY26", "quarterly"),
}

# SAHIs comparison: one column-block per company.
SAHI_CMP_BLOCKS = [
    ("niva-bupa", {"C": "FY23", "D": "FY24", "E": "FY25", "F": "FY26",
                   "H": "H1FY25", "I": "9MFY25", "J": "Q4FY25",
                   "K": "H1FY26", "L": "9MFY26", "M": "Q4FY26"}),
    ("star-health", {"P": "FY23", "Q": "FY24", "R": "FY25", "S": "FY26",
                     "U": "H1FY25", "V": "9MFY25", "W": "Q4FY25",
                     "X": "H1FY26", "Y": "9MFY26", "Z": "Q4FY26"}),
    ("care-health", {"AC": "FY23", "AD": "FY24", "AE": "FY25", "AF": "FY26",
                     "AH": "Q1FY25", "AI": "9MFY25", "AJ": "Q4FY25",
                     "AK": "Q1FY26", "AL": "9MFY26", "AM": "Q4FY26"}),
    ("manipalcigna", {"AP": "FY23", "AQ": "FY24", "AR": "FY25",
                      "AT": "Q1FY25", "AU": "Q1FY26"}),
    ("aditya-birla", {"AX": "FY23", "AY": "FY24", "AZ": "FY25",
                      "BB": "Q1FY25", "BC": "Q1FY26"}),
]

# SAHIs comparison: metric per row (row -> (metric, unit, source_key)).
SAHI_CMP_ROWS = {
    5:  ("retail_health_gwp", "INR_cr", "company_premium_quarterly"),
    6:  ("group_other_gwp", "INR_cr", "company_premium_quarterly"),
    7:  ("total_gwp", "INR_cr", "company_premium_quarterly"),
    8:  ("retail_health_market_share", "ratio", "company_financials"),
    9:  ("overall_health_market_share", "ratio", "company_financials"),
    11: ("nwp", "INR_cr", "company_financials"),
    12: ("nep", "INR_cr", "company_financials"),
    17: ("pat_igaap", "INR_cr", "company_financials"),
    19: ("claims_ratio_igaap", "ratio", "company_financials"),
    20: ("expense_ratio_igaap", "ratio", "company_financials"),
    21: ("combined_ratio_igaap", "ratio", "company_financials"),
    22: ("eom_igaap", "ratio", "company_financials"),
    25: ("pat_ifrs", "INR_cr", "company_financials"),
    27: ("claims_ratio_ifrs", "ratio", "company_financials"),
    28: ("expense_ratio_ifrs", "ratio", "company_financials"),
    32: ("solvency_ratio", "ratio", "company_financials"),
    33: ("net_worth", "INR_cr", "company_financials"),
    34: ("investment_aum", "INR_cr", "company_financials"),
    36: ("investment_yield", "ratio", "company_financials"),
}

# Comps: metric per column (col -> (metric, unit, source_key)).
COMPS_COLS = {
    "C": ("market_cap", "INR_cr", "market_cap"),
    "D": ("enterprise_value", "INR_cr", "market_cap"),
    "E": ("gwp", "INR_cr", "company_financials"),
    "F": ("net_worth_ifrs", "INR_cr", "company_financials"),
    "G": ("pat_ifrs", "INR_cr", "company_financials"),
    "H": ("net_worth_igaap", "INR_cr", "company_financials"),
    "I": ("pat_igaap", "INR_cr", "company_financials"),
    # Valuation MULTIPLES are "x" (42.5x, 4.3x, 1.8x), not percentages. ROE is a
    # return percentage computed as PAT / Net worth — a fraction (0.04) shown
    # ×100 as 4%, so it carries the "ratio" unit like every other percentage
    # metric in the workbook (claims / combined / expense ratios, market share).
    "J": ("price_to_gwp", "x", "market_cap"),
    "K": ("pe_igaap", "x", "market_cap"),
    "L": ("pb_igaap", "x", "market_cap"),
    "M": ("roe_igaap", "ratio", "company_financials"),
    "N": ("pe_ifrs", "x", "market_cap"),
    "O": ("pb_ifrs", "x", "market_cap"),
    "P": ("roe_ifrs", "ratio", "company_financials"),
    "Q": ("pe_3yr_avg", "x", "valuation_history"),
}
COMPS_ROWS = {4: "niva-bupa", 5: "star-health", 6: "icici-lombard", 7: "godigit"}

# Channel Mix: one column-block per company; FY19..9MFY26.
CHANNEL_AXIS_COLS = ["FY19", "FY20", "FY21", "FY22", "FY23", "FY24", "FY25", "9MFY26"]
CHANNEL_BLOCKS = [("niva-bupa", "C"), ("star-health", "K"), ("care-health", "S")]
CHANNEL_SECTIONS = {  # header_row -> (metric_prefix, unit, data_rows->channel)
    3:  ("channel_gwp_mix", "ratio"),
    12: ("avg_premium_per_policy", "INR_thousand"),
    21: ("commission_pct_gross", "ratio"),
}
CHANNEL_CHANNEL_ROWS = {  # offset from header row -> channel label
    1: "Banca", 2: "Brokers", 3: "Individual agents",
    4: "Corporate Agents - Others", 5: "Direct Business", 6: "Others", 7: "Total",
}


def entity_from_label(label: str) -> str:
    """Map a workbook row label to the repo's company_id where possible."""
    if not label:
        return ""
    key = label.strip().lower()
    table = {
        "star health": "star-health", "star": "star-health",
        "care": "care-health", "care health": "care-health",
        "niva bupa": "niva-bupa", "niva": "niva-bupa",
        "aditya birla": "aditya-birla",
        "manipal cigna": "manipalcigna", "manipalcigna": "manipalcigna",
        "reliance health": "reliance-health", "reliance": "reliance-general",
        "hdfc ergo": "hdfc-ergo", "hdfc ergo (incl. apollo munich)": "hdfc-ergo",
        "galaxy health": "galaxy-health", "narayana health": "narayana-health",
        "icici lombard": "icici-lombard", "bajaj allianz": "bajaj-general",
        "sbi gi": "sbi-general", "go digit": "godigit",
        "new india": "new-india", "national insurance": "national-insurance",
        "oriental insurance": "oriental-insurance", "united india": "united-india",
        "national": "national-insurance", "oriental": "oriental-insurance",
    }
    return table.get(key, "")


# ---------------------------------------------------------------------------
# Cell classification.
# ---------------------------------------------------------------------------
def classify(value, formula) -> str:
    if isinstance(formula, str) and formula.startswith("="):
        return "formula"
    if value is None and formula is None:
        return "empty"
    v = formula if formula is not None else value
    if isinstance(v, (int, float)):
        return "input"
    if isinstance(v, datetime):
        return "input_date"
    s = str(v).strip()
    if s == "":
        return "empty"
    # short non-numeric tokens that are really data placeholders
    if s.upper() in {"NA", "N/A", "NM", "N.A.", "N.M."}:
        return "input_na"
    return "text"


def main(template_path: Path) -> None:
    wb_v = openpyxl.load_workbook(template_path, data_only=True)
    wb_f = openpyxl.load_workbook(template_path, data_only=False)

    schema = {
        "_meta": {
            "artifact": "schema-map",
            "schema_version": "1.0.0",
            "phase": "1 - template reader",
            # Deterministic: keyed to template content (no wall-clock), so the
            # committed artifact only changes when the template actually changes.
            "template_sha256": hashlib.sha256(template_path.read_bytes()).hexdigest(),
            "template_file": template_path.name,
            "template_origin": "Built with the S&P Capital IQ Excel plug-in (see _CIQHiddenCacheSheet and CIQ() formulas in 'Comps'). This pipeline rebuilds those outputs from official public sources.",
            "source_policy": "official-first (IRDAI / GI Council / NSE-BSE / company public disclosures). Screener / Trendlyne / Investing are clearly-tagged, lower-confidence BACKUP only, used where no official equivalent exists. Login-free (Neha, 2026-06-05).",
            "cell_kinds": {
                "input": "literal value sourced from a fetcher - FILLABLE",
                "input_date": "literal date - FILLABLE",
                "input_na": "explicit NA/NM placeholder in the template",
                "formula": "Excel calculation - PRESERVE, do not overwrite",
                "text": "label / narrative text",
                "empty": "blank",
            },
            "provenance_contract": [
                "source_name", "source_url", "fetched_at", "period",
                "raw_value", "normalized_value", "transformation_used",
                "confidence", "source_status",
            ],
            "value_rules": [
                "Missing != zero. A missing source value is null + source_status=unavailable_publicly, never 0.",
                "Only 'input'/'input_date' cells are filled. 'formula' cells are preserved so the template recomputes.",
                "Period labels are honest (FY / Q / monthly / cumulative) per the column axis.",
            ],
        },
        "sources": SOURCES,
        "sheets": [],
    }

    for ws_v in wb_v.worksheets:
        name = ws_v.title
        ws_f = wb_f[name]
        kinds = {"input": 0, "input_date": 0, "input_na": 0, "formula": 0, "text": 0, "empty": 0}
        for r in range(1, ws_v.max_row + 1):
            for c in range(1, ws_v.max_column + 1):
                kinds[classify(ws_v.cell(r, c).value, ws_f.cell(r, c).value)] += 1

        sheet_obj = {
            "sheet": name,
            "dimensions": ws_v.dimensions,
            "max_row": ws_v.max_row,
            "max_col": ws_v.max_column,
            "merged_cells": [str(m) for m in ws_v.merged_cells.ranges],
            "hidden_columns": [c for c, d in ws_v.column_dimensions.items() if d.hidden],
            "grid_summary": kinds,
            "role": SHEET_ROLE.get(name, "unmapped"),
            "bindings": [],
        }
        builder = BUILDERS.get(name)
        if builder:
            sheet_obj["bindings"] = builder(ws_v, ws_f)
            sheet_obj["fillable_inputs"] = sum(1 for b in sheet_obj["bindings"] if b["fillable"])
        schema["sheets"].append(sheet_obj)

    OUT.write_text(json.dumps(schema, indent=2, ensure_ascii=False))

    total_bind = sum(len(s["bindings"]) for s in schema["sheets"])
    total_fill = sum(s.get("fillable_inputs", 0) for s in schema["sheets"])
    print(f"schema-map.json written -> {OUT}")
    print(f"  sheets: {len(schema['sheets'])}  bindings: {total_bind}  fillable inputs: {total_fill}")
    for s in schema["sheets"]:
        if s["bindings"]:
            print(f"    {s['sheet']:<26} role={s['role']:<22} "
                  f"bindings={len(s['bindings']):>4} fillable={s.get('fillable_inputs', 0):>4}")
        else:
            print(f"    {s['sheet']:<26} role={s['role']:<22} (grid only)")


# ---------------------------------------------------------------------------
# Binding builders (one per data sheet). Each returns a list of binding dicts.
# ---------------------------------------------------------------------------
def _binding(ws_v, ws_f, col, row, *, entity, metric, period, period_type, unit, source_key, section, conf="high", replace_formula=False):
    """Build one binding. ``replace_formula=True`` marks a cell that is a formula
    in the template only because it was fed by the S&P Capital IQ plug-in
    (``CIQ()``); since that paid plug-in is gone, the filler replaces it with a
    fetched value, so it counts as a fillable external input."""
    cell = f"{col}{row}"
    kind = classify(ws_v[cell].value, ws_f[cell].value)
    is_external_input = replace_formula and kind == "formula"
    src = SOURCES[source_key]
    return {
        "cell": cell,
        "section": section,
        "entity": entity,
        "metric": metric,
        "period": period,
        "period_type": period_type,
        "unit": unit,
        "cell_kind": kind,
        "is_external_input": is_external_input,
        "fillable": kind in ("input", "input_date", "input_na") or is_external_input,
        "template_value_preview": ws_v[cell].value if kind != "formula" else None,
        "source_key": source_key,
        "source_name": src["primary_source"],
        "source_status": src["status"],
        "binding_confidence": conf,
    }


def build_entity_rows(ws_v, ws_f, *, sections, axis, source_key, default_unit, period_type_override=None):
    """Orientation: entity per row (col B label), one metric per table, period across columns."""
    out = []
    for sec in sections:
        metric = sec["metric"]
        unit = sec.get("unit", default_unit)
        for row in range(sec["row_start"], sec["row_end"] + 1):
            label = ws_v[f"B{row}"].value
            if not label or ">>" in str(label):
                continue
            entity = sec.get("entity") or entity_from_label(str(label)) or str(label).strip()
            for col, (period, ptype) in axis.items():
                out.append(_binding(
                    ws_v, ws_f, col, row,
                    entity=entity, metric=metric, period=period,
                    period_type=period_type_override or ptype, unit=unit,
                    source_key=source_key, section=sec["name"],
                ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_industry_growth(ws_v, ws_f):
    axis = fy_axis("C", 15, 26)
    sections = [
        {"name": "GI industry premium by segment", "metric": "gi_segment_gross_premium",
         "row_start": 4, "row_end": 7, "unit": "INR_cr"},
        {"name": "Health premium by carrier type", "metric": "health_premium_by_carrier_type",
         "row_start": 11, "row_end": 14, "unit": "INR_cr"},
        {"name": "SAHI total health premium", "metric": "sahi_total_health_premium",
         "row_start": 18, "row_end": 27, "unit": "INR_cr"},
        {"name": "SAHI retail health premium", "metric": "sahi_retail_health_premium",
         "row_start": 42, "row_end": 49, "unit": "INR_cr"},
        {"name": "Retail health premium by insurer", "metric": "retail_health_premium",
         "row_start": 54, "row_end": 66, "unit": "INR_cr"},
    ]
    return build_entity_rows(ws_v, ws_f, sections=sections, axis=axis,
                             source_key="industry_premium", default_unit="INR_cr")


def build_fy26_gwp(ws_v, ws_f):
    sections = [
        {"name": "Total Health GWP", "metric": "total_health_gwp",
         "row_start": 5, "row_end": 24},
        {"name": "Retail Health GWP", "metric": "retail_health_gwp",
         "row_start": 29, "row_end": 48},
    ]
    return build_entity_rows(ws_v, ws_f, sections=sections, axis=GWP_FY_AXIS,
                             source_key="company_premium_quarterly", default_unit="INR_cr")


def build_q1_gwp(ws_v, ws_f):
    sections = [
        {"name": "Total Health GWP (monthly)", "metric": "total_health_gwp",
         "row_start": 5, "row_end": 24},
        {"name": "Retail Health GWP (monthly)", "metric": "retail_health_gwp",
         "row_start": 29, "row_end": 48},
    ]
    return build_entity_rows(ws_v, ws_f, sections=sections, axis=Q1_MONTHLY_AXIS,
                             source_key="company_premium_monthly", default_unit="INR_cr")


def build_sahis_comparison(ws_v, ws_f):
    out = []
    for entity, axis in SAHI_CMP_BLOCKS:
        for row, (metric, unit, source_key) in SAHI_CMP_ROWS.items():
            for col, period in axis.items():
                ptype = "annual" if period.startswith("FY") else "quarterly"
                out.append(_binding(
                    ws_v, ws_f, col, row,
                    entity=entity, metric=metric, period=period,
                    period_type=ptype, unit=unit, source_key=source_key,
                    section="SAHI detailed comparison",
                ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_comps(ws_v, ws_f):
    out = []
    # These columns were CIQ() plug-in formulas in the template; replace them.
    ciq_external = {"C", "D", "Q"}
    for row, entity in COMPS_ROWS.items():
        for col, (metric, unit, source_key) in COMPS_COLS.items():
            out.append(_binding(
                ws_v, ws_f, col, row,
                entity=entity, metric=metric, period="as_on_run_date",
                period_type="point_in_time", unit=unit, source_key=source_key,
                section="Valuation comps",
                conf="medium" if source_key == "valuation_history" else "high",
                replace_formula=col in ciq_external,
            ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_captable(ws_v, ws_f):
    out = []
    as_of = ws_v["C2"].value
    period = as_of.date().isoformat() if isinstance(as_of, datetime) else "as_of_filing"
    for row in range(4, 19):  # investors; row 19 = Total (formula)
        label = ws_v[f"B{row}"].value
        if not label:
            continue
        out.append(_binding(
            ws_v, ws_f, "D", row,
            entity="niva-bupa", metric=f"shareholding_shares::{str(label).strip()}",
            period=period, period_type="point_in_time", unit="shares",
            source_key="shareholding", section="Shareholding pattern",
        ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_channel_mix(ws_v, ws_f):
    out = []
    for entity, start_col in CHANNEL_BLOCKS:
        base = column_index_from_string(start_col)
        period_cols = {get_column_letter(base + i): CHANNEL_AXIS_COLS[i] for i in range(len(CHANNEL_AXIS_COLS))}
        for hdr_row, (metric, unit) in CHANNEL_SECTIONS.items():
            for off, channel in CHANNEL_CHANNEL_ROWS.items():
                row = hdr_row + off
                for col, period in period_cols.items():
                    ptype = "annual" if period.startswith("FY") else "quarterly_cumulative"
                    out.append(_binding(
                        ws_v, ws_f, col, row,
                        entity=entity, metric=f"{metric}::{channel}", period=period,
                        period_type=ptype, unit=unit, source_key="distribution",
                        section="Channel mix", conf="medium",
                    ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_hist_stock(ws_v, ws_f):
    out = []
    cols = {"C": ("close_price", "INR"), "D": ("traded_quantity", "shares"),
            "E": ("deliverable_quantity", "shares")}
    for row in range(3, 158):
        d = ws_v[f"B{row}"].value
        if not isinstance(d, datetime):
            continue
        period = d.date().isoformat()
        for col, (metric, unit) in cols.items():
            out.append(_binding(
                ws_v, ws_f, col, row,
                entity="niva-bupa", metric=metric, period=period,
                period_type="daily", unit=unit, source_key="market_quote",
                section="Historical stock movement",
            ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_analyst(ws_v, ws_f):
    """Each broker row -> reco/target. Captured as a record list (col-typed)."""
    out = []
    cols = {"E": ("recommendation", "text"), "G": ("price_at_reco", "INR"), "H": ("target_price", "INR")}
    current_entity = ""
    for row in range(4, 67):
        ent_label = ws_v[f"B{row}"].value
        if ent_label and entity_from_label(str(ent_label)):
            current_entity = entity_from_label(str(ent_label))
        broker = ws_v[f"C{row}"].value
        if not broker or str(broker).strip().lower() == "average":
            continue
        date = ws_v[f"D{row}"].value
        period = date.date().isoformat() if isinstance(date, datetime) else "undated"
        for col, (metric, unit) in cols.items():
            out.append(_binding(
                ws_v, ws_f, col, row,
                entity=current_entity, metric=f"analyst_{metric}::{str(broker).strip()}",
                period=period, period_type="event", unit=unit,
                source_key="analyst_coverage", section="Analyst coverage", conf="low",
            ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_mgmt_commentary(ws_v, ws_f):
    out = []
    company_cols = {"C": "niva-bupa", "D": "star-health", "E": "icici-lombard", "F": "godigit"}
    for row in range(2, 51):
        topic = ws_v[f"B{row}"].value
        if not topic:
            continue
        topic_s = str(topic).strip()
        if "Concall" in topic_s:  # quarter header row
            continue
        if topic_s not in {"Growth", "Profitability", "Competition", "Regulations", "Others / Highlights", "Others"}:
            continue
        for col, entity in company_cols.items():
            out.append(_binding(
                ws_v, ws_f, col, row,
                entity=entity, metric=f"commentary::{topic_s}", period="per_quarter_block",
                period_type="event", unit="text", source_key="management_commentary",
                section="Management commentary", conf="low",
            ))
    return [b for b in out if b["cell_kind"] != "empty"]


def build_sector_news(ws_v, ws_f):
    out = []
    for row in range(4, 35):
        subj = ws_v[f"E{row}"].value
        if not subj:
            continue
        date = ws_v[f"D{row}"].value
        period = date.date().isoformat() if isinstance(date, datetime) else "undated"
        out.append(_binding(
            ws_v, ws_f, "F", row,
            entity="sector", metric="news_summary", period=period,
            period_type="event", unit="text", source_key="sector_news",
            section="Key sectoral updates", conf="low",
        ))
    return [b for b in out if b["cell_kind"] != "empty"]


SHEET_ROLE = {
    "Industry Growth": "industry_premium",
    "FY26 GWP": "company_premium_quarterly",
    "Q1'26 GWP": "company_premium_monthly",
    "SAHIs comparison": "company_financials",
    "Management commentary ": "management_commentary",
    "Analyst coverage": "analyst_coverage",
    "Comps": "valuation",
    "Captable": "shareholding",
    "Channel Mix": "distribution",
    "Key sectoral updates": "sector_news",
    "_CIQHiddenCacheSheet": "ignore_plugin_cache",
    "Historical Stock Movement": "market_quote",
}

BUILDERS = {
    "Industry Growth": build_industry_growth,
    "FY26 GWP": build_fy26_gwp,
    "Q1'26 GWP": build_q1_gwp,
    "SAHIs comparison": build_sahis_comparison,
    "Comps": build_comps,
    "Captable": build_captable,
    "Channel Mix": build_channel_mix,
    "Historical Stock Movement": build_hist_stock,
    "Analyst coverage": build_analyst,
    "Management commentary ": build_mgmt_commentary,
    "Key sectoral updates": build_sector_news,
}


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    if not path.exists():
        sys.exit(f"Template not found: {path}\nPass the .xlsx path as the first argument.")
    main(path)
