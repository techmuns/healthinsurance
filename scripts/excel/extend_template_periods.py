#!/usr/bin/env python3
"""
Template PERIOD auto-extension (standing instruction, Neha, 2026-06-11).

When the data layer holds a period the template has no column for (a new
fiscal year on Industry Growth, a new FY quarter-group on the GWP tab, a new
Q1 month-group on the monthly tab), this script APPENDS the column group to
templates/niva-bupa-portfolio-review.xlsx in the established format, so the
next schema-map build binds it and the grid fills it on the same run — no
human in the loop. "Even if I see this whole dashboard after 2 years, I see
the best, highest-quality, live-updated data."

Safety rules:
  * Extend ONLY when real data for the new period exists in the snapshots —
    never a speculative empty column.
  * APPEND into virgin columns after the sheet's used range — existing cells,
    formulas and the analysis blocks (Mix %, YoY, CAGR) are never shifted or
    rewritten.
  * New cells are INPUTS (no formula cloning): every value they show comes
    from the sourced data layer, which already provides standalone quarters
    and single months by exact arithmetic over printed cumulatives.
  * Number formats / fonts / borders are copied from the equivalent cell of
    the donor (latest existing) period group, so the look stays identical.
  * Period-column replication only — never new rows, metrics or sections.
"""
from __future__ import annotations

import json
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
TEMPLATE = REPO / "templates" / "niva-bupa-portfolio-review.xlsx"
SNAP = REPO / "src" / "data" / "snapshots"


def load_periods(name: str, field: str) -> set[str]:
    try:
        rows = json.loads((SNAP / f"{name}.json").read_text()).get("data", [])
        return {str(r.get(field)) for r in rows if r.get(field)}
    except Exception:
        return set()


def copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)


def append_group(ws, donor_cols: list[int], headers: dict[int, dict[int, object]],
                 banner: str | None, first_data_row: int, last_row: int) -> list[str]:
    """Clone the donor columns' look into the first free columns; write the
    given header texts; leave every data cell empty (inputs for the grid)."""
    start = ws.max_column + 1
    added = []
    for offset, donor in enumerate(donor_cols):
        col = start + offset
        ws.column_dimensions[get_column_letter(col)].width = (
            ws.column_dimensions[get_column_letter(donor)].width)
        for row in range(1, last_row + 1):
            d = ws.cell(row=row, column=donor)
            c = ws.cell(row=row, column=col)
            copy_style(d, c)
            if row in headers.get(offset, {}):
                c.value = headers[offset][row]
        added.append(get_column_letter(col))
    if banner is not None:
        ws.cell(row=2, column=start).value = banner
    return added


def next_fy(fy: str) -> str:
    return f"FY{int(fy[2:]) + 1:02d}"


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    changed: list[str] = []

    portfolio_fys = load_periods("gic-health-portfolio", "fiscal_year")
    quarterly = load_periods("gic-health-quarterly", "period")
    monthly = load_periods("gic-health-monthly", "period")

    # ── Industry Growth: one column per fiscal year ─────────────────────────
    ws = wb["Industry Growth"]
    # Current last FY column: header row 3 runs 15..N as a +1 formula chain
    # anchored at C3=15; the chain length gives the last covered FY.
    fy_cols = 0
    col = 3
    while ws.cell(row=3, column=col).value is not None and (
            isinstance(ws.cell(row=3, column=col).value, (int, float))
            or str(ws.cell(row=3, column=col).value).startswith("=")):
        fy_cols += 1
        col += 1
    last_fy = 14 + fy_cols  # C3 = 15
    nxt = f"FY{last_fy + 1 - 2000 if last_fy + 1 >= 2000 else last_fy + 1:02d}"
    if nxt in portfolio_fys:
        donor = 2 + fy_cols  # the last FY column index
        # Header rows: every section repeats its own year-header row — find
        # all rows whose donor-column cell is part of the +1 chain or a year
        # number, and bump them.
        headers: dict[int, dict[int, object]] = {0: {}}
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=donor).value
            if isinstance(v, (int, float)) and 15 <= v <= 99:
                headers[0][row] = int(v) + 1
            elif isinstance(v, str) and v.startswith("=") and "+1" in v:
                headers[0][row] = last_fy + 1
        added = append_group(ws, [donor], headers, None, 4, ws.max_row)
        changed.append(f"Industry Growth + {nxt} column ({added[0]})")

    # ── FY26 GWP: a 6-column group per fiscal year ──────────────────────────
    ws = wb["FY26 GWP"]
    hdr_row = 3
    existing = {str(ws.cell(row=hdr_row, column=c).value) for c in range(1, ws.max_column + 1)}
    gwp_fys = sorted({p[-4:] for p in existing if p.startswith("Q1FY")})
    if gwp_fys:
        last = gwp_fys[-1]  # e.g. 'FY26'
        nxt = next_fy(last)
        has_data = any(p.endswith(nxt) for p in quarterly) or nxt in portfolio_fys
        if has_data and f"Q1{nxt}" not in existing:
            donors = [c for c in range(1, ws.max_column + 1)
                      if str(ws.cell(row=hdr_row, column=c).value).endswith(last)
                      and ("FY" in str(ws.cell(row=hdr_row, column=c).value))]
            labels = [str(ws.cell(row=hdr_row, column=c).value).replace(last, nxt) for c in donors]
            headers = {i: {hdr_row: labels[i]} for i in range(len(donors))}
            added = append_group(ws, donors, headers, nxt, 4, ws.max_row)
            changed.append(f"FY26 GWP + {nxt} group ({added[0]}..{added[-1]}: {', '.join(labels)})")

    # ── Q1'26 GWP: a 4-column month group per fiscal year ───────────────────
    ws = wb["Q1'26 GWP"]
    hdr_row = 3
    existing = {str(ws.cell(row=hdr_row, column=c).value) for c in range(1, ws.max_column + 1)}
    q1_fys = sorted({p[-4:] for p in existing if p.startswith("Q1FY")})
    if q1_fys:
        last = q1_fys[-1]
        nxt = next_fy(last)
        # months of FY27 are Apr'26..Jun'26 → calendar yy = FY - 1
        cal = int(nxt[2:]) - 1
        wanted = [f"Apr'{cal}", f"May'{cal}", f"Jun'{cal}", f"Q1{nxt}"]
        has_data = any(p in monthly for p in (f"Apr-{nxt}", f"May-{nxt}", f"Jun-{nxt}")) \
            or f"Q1{nxt}" in quarterly
        if has_data and f"Q1{nxt}" not in existing:
            # Donor = the CURRENT-FY group: the FIRST occurrence of each label
            # (the YoY block on the right repeats the month labels — skip it by
            # taking the first match per label, scanning from the left).
            prev_cal = cal - 1
            donor_labels = [f"Apr'{prev_cal}", f"May'{prev_cal}", f"Jun'{prev_cal}", f"Q1{last}"]
            donors = []
            for lbl in donor_labels:
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(row=hdr_row, column=c).value) == lbl:
                        donors.append(c)
                        break
            if len(donors) == 4:
                headers = {i: {hdr_row: wanted[i]} for i in range(4)}
                added = append_group(ws, donors, headers, nxt, 4, ws.max_row)
                changed.append(f"Q1'26 GWP + {nxt} month group ({added[0]}..{added[-1]}: {', '.join(wanted)})")

    if changed:
        wb.save(TEMPLATE)
        for c in changed:
            print(f"extended: {c}")
    else:
        print("no new periods with data — template unchanged")


if __name__ == "__main__":
    main()
