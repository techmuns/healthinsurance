#!/usr/bin/env python3
"""
Phase 8 - Extracted Data Audit index builder.

Projects the EXISTING source-backed pipeline artifacts into one compact,
browser-loadable JSON that powers the dashboard's **Extracted Data Audit** tab
(a QA surface). It does NOT re-source or re-derive any number - it only SELECTS
fields from artifacts the rest of the pipeline already produced:

    schema-map.json                         (Phase 1 - the cell-level contract)
    data/processed/excel-values.json        (Phase 2 - the normalized value store)
    data/processed/excel-held-back.json     (Phase 2 - extracted-but-withheld)
    src/data/snapshots/company-filings-snapshot.json  (blocked / parser-gated filings)

Output:
    src/data/snapshots/extracted-data-audit.json

The join itself (binding x value -> Fetched / Missing / Parser issue / ...) lives
ONCE, on the TypeScript side (src/lib/extractedDataAudit.ts), so the audit tab
reuses the dashboard's normalized data instead of duplicating data logic. This
script only distils the inputs so the bundle stays lean (we never import the
1.8 MB schema-map or the raw stores into the browser).

Honesty model is inherited verbatim from the pipeline: missing != zero, official
sources first, template treated as layout-only, period/basis labels honest.

Usage:
    python3 scripts/excel/build_audit_index.py
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
SCHEMA = REPO / "schema-map.json"
VALUES = REPO / "data" / "processed" / "excel-values.json"
HELD_BACK = REPO / "data" / "processed" / "excel-held-back.json"
FILINGS = REPO / "src" / "data" / "snapshots" / "company-filings-snapshot.json"
TEMPLATE = REPO / "templates" / "niva-bupa-portfolio-review.xlsx"
OUT = REPO / "src" / "data" / "snapshots" / "extracted-data-audit.json"

# Max referenced cells we record per formula (keeps a big SUM range bounded).
MAX_FORMULA_INPUTS = 24

# How deep the formula evaluator will chase a chain of cell-to-cell references
# before giving up. A ratio cell (e.g. Comps P/E = Market Cap / PAT) divides
# intermediate cells that THEMSELVES pull a number from another sheet, so the
# resolver has to follow two or three hops to reach a source-backed value. The
# cap guards against a cyclic reference; the real chains here are 2–3 deep.
MAX_RESOLVE_DEPTH = 8

# Cell kinds we surface as audit rows. `text` / `empty` are pure layout labels
# (no value contract) and are skipped. Everything else is shown — including
# `formula` cells (computed in-sheet) so the coverage check sees EVERY cell the
# template defines (e.g. the combined-ratio cells some sheets derive rather than
# fetch). Formula cells are tagged `computed` at read time; several already carry
# a directly-fetched value in the store, which the tab surfaces.
AUDIT_CELL_KINDS = {"input", "input_date", "input_na", "formula"}

# Roles that are not part of the data contract (the S&P Capital IQ plugin cache).
SKIP_ROLES = {"ignore_plugin_cache"}

# Binding fields kept per cell (compact). We DROP the original template number
# (layout-only, never re-used), the per-cell `source_name` (derived at read time
# from `source_key` -> sources registry), and `binding_confidence` (almost
# always "high"; value-level confidence is what QA cares about).
BINDING_FIELDS = (
    "cell", "section", "entity", "metric", "period", "period_type",
    "unit", "cell_kind", "fillable", "source_key", "source_status",
)

# Value-store fields kept per entity::metric::period entry.
VALUE_FIELDS = (
    "entity", "metric", "period", "unit", "raw_value", "normalized_value",
    "transformation_used", "source_name", "source_url", "source_file",
    "fetched_at", "filing_date", "confidence", "source_status", "source_layer",
    "priority_rank", "document_type", "document_title", "extraction_status",
    "conflict_status", "basis_note", "eligible_for_excel",
)

HELD_FIELDS = (
    "company_id", "metric", "filing_period", "raw_value", "normalized_value",
    "unit", "document_type", "document_title", "filing_date", "source_url",
    "source_file", "confidence", "hold_reason", "note",
)

FILING_FIELDS = (
    "company_id", "metric", "filing_period", "document_type", "document_title",
    "raw_value", "normalized_value", "unit", "filing_date", "source_url",
    "source_file", "extraction_status", "sanity_status", "sanity_reason",
    "parser_notes", "suggested_manual_fallback",
)


def load(path: Path, default):
    return json.loads(path.read_text()) if path.exists() else default


def pick(d: dict, fields) -> dict:
    """Copy only `fields` that are present (keeps the payload compact)."""
    return {k: d[k] for k in fields if k in d and d[k] is not None}


def latest_iso(*candidates) -> str | None:
    vals = [c for c in candidates if isinstance(c, str) and c]
    return max(vals) if vals else None


# ---------------------------------------------------------------------------
#  Formula resolver — reads the committed template and, for each computed
#  (formula) cell, recovers: the raw Excel formula, the cells it references
#  (resolved to their metric / period / row label so a reviewer can see WHERE
#  each number comes from), and a readable "calculation in words". Guarded so
#  the build still succeeds if openpyxl/the template isn't present (formula
#  detail is simply omitted then).
# ---------------------------------------------------------------------------

def build_formula_resolver(schema, store):
    try:
        import ast
        import operator
        import re
        import openpyxl
        from openpyxl.formula.tokenizer import Tokenizer
        from openpyxl.utils import range_boundaries, get_column_letter
    except Exception:
        return None
    if not TEMPLATE.exists():
        return None
    try:
        wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    except Exception:
        return None

    # cell -> binding (entity / metric / period) per sheet, from the schema map.
    cellmap = {sh.get("sheet"): {b.get("cell"): b for b in (sh.get("bindings") or [])}
               for sh in schema.get("sheets", [])}
    # row -> human label per sheet, read from the template's leftmost text column.
    rowlabel = {}
    for ws in wb.worksheets:
        rl = {}
        for r in range(1, (ws.max_row or 0) + 1):
            for c in range(1, min(ws.max_column or 1, 5) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip() and not v.startswith("="):
                    rl[r] = v.strip()
                    break
        rowlabel[ws.title] = rl

    coord_re = re.compile(r"([A-Z]+)(\d+)")

    def split_ref(value):
        external = False
        sheet = None
        ref = value
        if "!" in value:
            sp, ref = value.rsplit("!", 1)
            sp = sp.strip()
            if sp.startswith("["):
                external = True
            else:
                sheet = sp.strip("'")
        return sheet, ref.replace("$", ""), external

    def expand(ref):
        if ":" in ref:
            try:
                c0, r0, c1, r1 = range_boundaries(ref)
            except Exception:
                return [ref]
            return [f"{get_column_letter(c)}{r}" for r in range(r0, r1 + 1) for c in range(c0, c1 + 1)]
        return [ref]

    def describe(sheet, coord, cur_sheet):
        s = sheet or cur_sheet
        b = cellmap.get(s, {}).get(coord) or {}
        m = coord_re.match(coord)
        row = int(m.group(2)) if m else None
        label = rowlabel.get(s, {}).get(row) or b.get("metric") or coord
        out = {"ref": coord, "label": label}
        if sheet and sheet != cur_sheet:
            out["sheet"] = sheet
        if b.get("period"):
            out["period"] = b["period"]
        if b.get("entity"):
            out["entity"] = b["entity"]
        if b.get("metric"):
            out["metric"] = b["metric"]
        return out

    # ---- value-from-our-data: evaluate a formula using the source-backed store,
    #      but ONLY when every referenced cell has a value (missing != guessed). ----
    def store_value(sheet, coord, cur_sheet, depth=0):
        s = sheet or cur_sheet
        b = cellmap.get(s, {}).get(coord) or {}
        e, m, p = b.get("entity"), b.get("metric"), b.get("period")
        if e and m and p:
            v = store.get(f"{e}::{m}::{p}")
            if v and v.get("normalized_value") is not None:
                return v["normalized_value"]
        # No direct store value for this cell. If the cell is ITSELF an in-sheet
        # formula, follow it: the Comps multiples (P/GWP, P/E, P/B, ROE) divide
        # intermediate cells (GWP, PAT, Net worth) that each pull their number
        # from the 'SAHIs comparison' grid via a cross-sheet reference. Chasing
        # those references reproduces the template's own calculation rather than
        # fetching the ratio from anywhere. External (Capital IQ) refs resolve to
        # None and so leave the cell honestly blank. Recursion is depth-capped.
        if depth < MAX_RESOLVE_DEPTH:
            return compute(s, coord, depth + 1)
        return None

    _OPS = {ast.Add: operator.add, ast.Sub: operator.sub, ast.Mult: operator.mul,
            ast.Div: operator.truediv, ast.Pow: operator.pow,
            ast.USub: operator.neg, ast.UAdd: operator.pos}

    def _seval(node):
        if isinstance(node, ast.BinOp):
            return _OPS[type(node.op)](_seval(node.left), _seval(node.right))
        if isinstance(node, ast.UnaryOp):
            return _OPS[type(node.op)](_seval(node.operand))
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        raise ValueError("unsupported expression")

    sum_re = re.compile(r"(SUM|AVERAGE)\(([^()]+)\)", re.I)
    ref_re = re.compile(r"(?:'[^']+'|\[\d+\][^!']*'?)?!?\$?[A-Z]{1,3}\$?\d+")

    def compute(cur_sheet, coord, depth=0):
        ws = wb[cur_sheet] if cur_sheet in wb.sheetnames else None
        if ws is None:
            return None
        raw = ws[coord].value
        if not (isinstance(raw, str) and raw.startswith("=")):
            return None
        if "IFERROR" in raw.upper():
            return None            # error-guarded formulas: leave to the reviewer
        bad = [False]
        expr = raw[1:]

        def repl_sum(m):
            fn = m.group(1).upper()
            vals = []
            for part in m.group(2).split(","):
                part = part.strip()
                for c in (expand(part) if ":" in part else [part.replace("$", "")]):
                    sheet, ref, ext = split_ref(c)
                    v = None if ext else store_value(sheet, ref, cur_sheet, depth)
                    if v is None:
                        bad[0] = True
                        return "0"
                    vals.append(v)
            if not vals:
                bad[0] = True
                return "0"
            joined = "+".join(repr(v) for v in vals)
            return f"({joined})" if fn == "SUM" else f"(({joined})/{len(vals)})"

        def repl_ref(m):
            sheet, ref, ext = split_ref(m.group(0))
            v = None if ext else store_value(sheet, ref, cur_sheet, depth)
            if v is None:
                bad[0] = True
                return "0"
            return repr(v)

        expr = sum_re.sub(repl_sum, expr)
        expr = ref_re.sub(repl_ref, expr)
        if bad[0]:
            return None
        expr = expr.replace("^", "**")
        expr = re.sub(r"(\d+(?:\.\d+)?)\s*%", r"(\1/100)", expr)
        try:
            return round(_seval(ast.parse(expr, mode="eval").body), 6)
        except Exception:
            return None

    def resolve(cur_sheet, coord):
        ws = wb[cur_sheet] if cur_sheet in wb.sheetnames else None
        if ws is None:
            return None
        raw = ws[coord].value
        if not (isinstance(raw, str) and raw.startswith("=")):
            return None
        try:
            toks = Tokenizer(raw).items
        except Exception:
            return {"formula": raw, "value": compute(cur_sheet, coord)}
        inputs, seen, words = [], set(), []
        for t in toks:
            if t.type == "OPERAND" and t.subtype == "RANGE":
                sheet, ref, external = split_ref(t.value)
                if external:
                    words.append("[external workbook]")
                    continue
                coords = expand(ref)
                labs = [describe(sheet, c, cur_sheet) for c in coords]
                for lab in labs:
                    key = (lab.get("sheet"), lab["ref"])
                    if key not in seen and len(inputs) < MAX_FORMULA_INPUTS:
                        seen.add(key)
                        inputs.append(lab)
                words.append(f"[{labs[0]['label']}]" if len(labs) == 1
                             else f"[{labs[0]['label']} … {labs[-1]['label']}]")
            elif t.type == "FUNC":
                words.append(t.value)            # e.g. "SUM(" or ")"
            elif t.type == "PAREN":
                words.append(t.value)
            elif t.type == "SEP":
                words.append(",")
            elif t.type.startswith("OPERATOR"):
                words.append(t.value)
            elif t.type == "OPERAND" and t.subtype == "NUMBER":
                words.append(t.value)
            elif t.type == "OPERAND" and t.subtype == "TEXT":
                words.append(f'"{t.value}"')
        calc = " ".join(w for w in words if w).strip()
        info = {"formula": raw}
        if inputs:
            info["inputs"] = inputs
        if calc:
            info["calc"] = calc
        value = compute(cur_sheet, coord)
        if value is not None:
            info["value"] = value
        return info

    return resolve


def _add_full_grid(sheets) -> int:
    """Add every (entity x metric x period) cell of the SAHIs comparison grid
    that the schema map dropped for being blank, as a fillable input cell.
    Returns how many cells were added. Safe no-op if the grid defs can't load."""
    try:
        import sys as _sys
        _sys.path.insert(0, str(Path(__file__).resolve().parent))
        from build_schema_map import SAHI_CMP_BLOCKS, SAHI_CMP_ROWS
    except Exception:
        return 0
    sheet = next((s for s in sheets if s["sheet"] == "SAHIs comparison"), None)
    if not sheet:
        return 0
    have = {c["cell"] for c in sheet["cells"]}
    added = 0
    for entity, axis in SAHI_CMP_BLOCKS:
        for row, (metric, unit, source_key) in SAHI_CMP_ROWS.items():
            for col, period in axis.items():
                ref = f"{col}{row}"
                if ref in have:
                    continue
                sheet["cells"].append({
                    "cell": ref,
                    "section": "SAHI detailed comparison",
                    "entity": entity,
                    "metric": metric,
                    "period": period,
                    "period_type": "annual" if period.startswith("FY") else "quarterly",
                    "unit": unit,
                    "cell_kind": "input",
                    "fillable": True,
                    "source_key": source_key,
                    "source_status": "available",
                })
                have.add(ref)
                added += 1
    return added


def main() -> None:
    schema = load(SCHEMA, {"sheets": [], "_meta": {}, "sources": {}})
    store = load(VALUES, {})
    held = load(HELD_BACK, {"data": []}).get("data", [])
    filings = load(FILINGS, {"data": []}).get("data", [])

    resolve_formula = build_formula_resolver(schema, store)

    # --- Sheets -> trimmed audit cells -----------------------------------
    sheets = []
    total_cells = 0
    total_computed = 0
    formula_resolved = 0
    for sh in schema.get("sheets", []):
        role = sh.get("role")
        if role in SKIP_ROLES:
            continue
        cells = []
        computed = 0
        for b in sh.get("bindings", []) or []:
            kind = b.get("cell_kind")
            if kind not in AUDIT_CELL_KINDS:
                continue
            cell = pick(b, BINDING_FIELDS)
            if kind == "formula":
                computed += 1
                if resolve_formula:
                    info = resolve_formula(sh.get("sheet"), b.get("cell"))
                    if info:
                        cell["formula"] = info.get("formula")
                        if info.get("calc"):
                            cell["calc"] = info["calc"]
                        if info.get("inputs"):
                            cell["inputs"] = info["inputs"]
                        if info.get("value") is not None:
                            cell["calculated_value"] = info["value"]
                        formula_resolved += 1
            cells.append(cell)
        if not cells and not computed:
            continue
        total_cells += len(cells)
        total_computed += computed
        sheets.append({
            "sheet": sh.get("sheet"),
            "role": role,
            "dimensions": sh.get("dimensions"),
            # How many of this sheet's cells are computed-in-Excel formulas (also
            # included in `cells`, tagged `computed` by the reader).
            "computed_cells": computed,
            "cells": cells,
        })

    # --- Full-grid coverage for the SAHIs comparison sheet ----------------
    # The schema builder DROPS cells that are blank in the template, so a blank
    # grid cell never reaches coverage. A coverage check needs every cell in the
    # grid (entity x metric x period) — blanks included, so a reviewer can see
    # "we don't have this yet" and our extracted values land in their real cell.
    total_cells += _add_full_grid(sheets)

    # --- IRDAI-blocked industry-history cells -----------------------------
    # The industry GI-segment premium history (Health / Motor / Others, the
    # pre-FY24 years) is published in the IRDAI Handbook, but IRDAI blocks
    # automated downloads and every proxy corrupts the files in transit. Mark
    # those specific empty cells "web_blocked" so they read "IRDAI web blocked"
    # instead of a generic "not reachable". Only cells with no source value.
    for sh in sheets:
        for cell in sh["cells"]:
            if cell.get("metric") != "gi_segment_gross_premium":
                continue
            k = f'{cell.get("entity")}::{cell.get("metric")}::{cell.get("period")}'
            v = store.get(k)
            if not (v and v.get("normalized_value") is not None):
                cell["source_status"] = "web_blocked"

    # --- Genuinely not-applicable cells -----------------------------------
    # Rows for insurers that did not exist in that period (not yet licensed /
    # merged away / exited), curated with reasons in
    # data/source-map/not-applicable-cells.json. Shown as "not applicable" —
    # never a fake zero, never a red "missing". Only applies while the cell
    # has no sourced value, so real data always wins if it ever appears.
    try:
        na_cells = json.loads((REPO / "data" / "source-map" / "not-applicable-cells.json").read_text()).get("cells", {})
    except Exception:
        na_cells = {}
    if na_cells:
        for sh in sheets:
            for cell in sh["cells"]:
                k = f'{cell.get("entity")}::{cell.get("metric")}::{cell.get("period")}'
                reason = na_cells.get(k)
                if not reason:
                    continue
                v = store.get(k)
                if not (v and v.get("normalized_value") is not None):
                    cell["source_status"] = "not_applicable"
                    cell["na_reason"] = reason

    # --- 'Not found in PPT' cells (grey) ----------------------------------
    # Cells whose company investor presentations / annual reports were swept
    # page-by-page (scripts/ingest/ppt-sweep.ts + reviewed transcription) and
    # genuinely do not print the number. Curated with per-cell reasons in
    # data/source-map/ppt-search-results.json (Neha, 2026-06-11: tag these
    # 'not found in ppt' and colour the cell grey). Only applies while the
    # cell has no sourced value — a statutory filing can still fill it later.
    try:
        ppt_neg = json.loads((REPO / "data" / "source-map" / "ppt-search-results.json").read_text()).get("cells", {})
    except Exception:
        ppt_neg = {}
    if ppt_neg:
        for sh in sheets:
            for cell in sh["cells"]:
                k = f'{cell.get("entity")}::{cell.get("metric")}::{cell.get("period")}'
                entry = ppt_neg.get(k)
                if not entry:
                    continue
                v = store.get(k)
                if not (v and v.get("normalized_value") is not None):
                    cell["source_status"] = "not_in_ppt"
                    cell["na_reason"] = entry.get("reason") if isinstance(entry, dict) else str(entry)

    # --- Value store (trimmed) -------------------------------------------
    values = {key: pick(entry, VALUE_FIELDS) for key, entry in store.items()}

    # --- Held-back (extracted but withheld) ------------------------------
    held_back = [pick(h, HELD_FIELDS) for h in held]

    # --- Blocked filings (parser-gated / not eligible for the template) --
    blocked_filings = [
        pick(r, FILING_FIELDS) for r in filings if not r.get("eligible_for_excel")
    ]

    # --- Source registry (role -> intended official source) --------------
    sources = {
        k: {
            "primary_source": v.get("primary_source"),
            "primary_url": v.get("primary_url"),
            "status": v.get("status"),
        }
        for k, v in (schema.get("sources") or {}).items()
    }

    sm_meta = schema.get("_meta", {})
    out = {
        "_meta": {
            "artifact": "extracted-data-audit",
            "schema_version": "1.0.0",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "description": (
                "QA index for the Extracted Data Audit tab. A compact projection "
                "of schema-map.json (cell contract) + excel-values.json (normalized "
                "value store) + held-back + blocked-filings. The binding x value join "
                "(Fetched / Missing / Parser issue / Source unavailable / Manual "
                "override) is computed at read time in src/lib/extractedDataAudit.ts "
                "so no data logic is duplicated."
            ),
            "template_file": sm_meta.get("template_file"),
            "template_sha256": sm_meta.get("template_sha256"),
            "source_policy": sm_meta.get("source_policy"),
            "provenance_contract": sm_meta.get("provenance_contract"),
            "value_rules": sm_meta.get("value_rules"),
            "last_updated": latest_iso(
                sm_meta.get("generated_at"),
                load(HELD_BACK, {}).get("_meta", {}).get("last_updated"),
                load(FILINGS, {}).get("_meta", {}).get("last_updated"),
            ),
            "counts": {
                "sheets": len(sheets),
                "audit_cells": total_cells,
                "computed_cells": total_computed,
                "value_store_entries": len(values),
                "held_back": len(held_back),
                "blocked_filings": len(blocked_filings),
            },
        },
        "sources": sources,
        "sheets": sheets,
        "values": values,
        "held_back": held_back,
        "blocked_filings": blocked_filings,
    }

    # If formulas couldn't be resolved (openpyxl/template absent — e.g. the
    # Cloudflare build) but a committed index already carries resolved formula
    # values, KEEP it rather than overwrite with empty formula cells. This lets a
    # locally-resolved index (P/E, P/B, ROE, …) survive a deploy that lacks
    # openpyxl. When openpyxl IS present we always rewrite (fresh data wins).
    if resolve_formula is None and total_computed and OUT.exists():
        try:
            prior = json.loads(OUT.read_text())
            prior_resolved = sum(1 for s in prior.get("sheets", []) for c in s.get("cells", []) if c.get("formula"))
        except Exception:
            prior_resolved = 0
        if prior_resolved:
            print(f"openpyxl unavailable — preserving committed index with {prior_resolved} resolved formula(s); not overwriting.")
            return

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=1))
    size_kb = OUT.stat().st_size / 1024
    print(f"extracted-data-audit.json written -> {OUT} ({size_kb:.0f} KB)")
    print(f"  sheets: {len(sheets)} | audit cells: {total_cells} | computed: {total_computed} (formulas resolved: {formula_resolved})")
    print(f"  value-store entries: {len(values)} | held-back: {len(held_back)} | blocked filings: {len(blocked_filings)}")
    if not formula_resolved and total_computed:
        print("  note: formula detail omitted (openpyxl/template unavailable) — coverage unaffected")


if __name__ == "__main__":
    main()
